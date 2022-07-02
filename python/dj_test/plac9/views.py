from django.shortcuts import render
from plac8.models import Customer,Recieve,Item
import openpyxl


def index(request):
    cusms=Customer.objects.all()
    return render(request,"plac9/index.html",{"cusms":cusms})


def upload(request):
    context = {}
    if request.method == 'POST' and request.FILES['excel']:
        excel = request.FILES['excel']
        #excelの読み込み
        wb = openpyxl.load_workbook(excel)

        ###################
        #顧客レポート
        ws = wb["顧客レポート"]

        all_data=[]
        for row in ws.iter_rows(min_row=2):
            line=[]
            for cell in row:
                line.append(cell.value)
            all_data.append(line)

        for i in all_data:
            Customer.objects.update_or_create(
                cus_id=i[0],
                defaults={
                    "cus_id":i[0],
                    "sei":i[1],
                    "mei":i[2],
                    "sei_kana":i[3],
                    "mei_kana":i[4],
                    "mail1":i[5],
                    "mail2":i[6],
                    "mail3":i[7],
                    "yubin":i[8],
                    "pref":i[9],
                    "city":i[10],
                    "banchi":i[11],
                    "build":i[12],
                    "com":i[13],
                    "busho":i[14],
                    "tel":i[15],
                    "mob":i[16],
                    "fax":i[17],
                    "toroku":i[18],
                    "kensu":i[19],
                    "money":i[20],
                }            
            )
        
        ###################
        #見積受注
        ws = wb["見積受注"]

        all_data=[]
        for row in ws.iter_rows(min_row=2):
            line=[]
            for cell in row:
                line.append(cell.value)
            all_data.append(line)

        for i in all_data:
            Recieve.objects.update_or_create(
                rec_id=i[0],               
                defaults={
                    "rec_id":i[0],
                    "rec_no":i[1],
                    "rec_ver":i[2],
                    "status":i[3],
                    "mitsu_day":str(i[4])[:10],
                    "rec_day":str(i[5])[:10],
                    "eigyou_sei":i[6],
                    "eigyou_mei":i[7],
                    "eigyou_busho":i[8],
                    "rec_cus_id":Customer.objects.get(cus_id=i[9]),
                    "keiro":i[10],
                    "mitsu_money":i[11],
                }                
            )


        ###################
        #見積商品
        ws = wb["見積商品"]

        all_data=[]
        for row in ws.iter_rows(min_row=2):
            line=[]
            for cell in row:
                line.append(cell.value)
            all_data.append(line)

        for i in all_data:
            Item.objects.create(
                item_rec_id=Recieve.objects.get(rec_id=i[0]),
                item_name=i[1],           
            )


    return render(request, 'plac9/index.html',{"test":"登録しました"})

